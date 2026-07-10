"""Additional THG edge materializers from on-disk transport sources."""

from __future__ import annotations

import json
import re
import zipfile
from pathlib import Path
from typing import Any

import duckdb

from catalog.config import ROOT, UNWRAPPED
from catalog.gaps import USER_AGENT
from catalog.path_resolver import resolve_family_file
from catalog.signals._common import CSV_OPTS
from catalog.source_truth import require_path, resolve_path

# HS chapter → SCTG2 coarse map for Census IMDB
_HS_SCTG = (
    (r"^27", "17"),  # mineral fuels / petroleum
    (r"^2701", "19"),  # coal
    (r"^2601", "19"),  # iron ore → coal family proxy
    (r"^1001|^1002|^1003|^1004|^1005|^1006|^1007|^1008", "02"),  # grain
    (r"^72|^73", "11"),  # iron/steel
    (r"^28|^29|^38|^39", "18"),  # chemicals / plastics
)


def _hs_to_sctg(hs: str) -> str:
    hs = re.sub(r"[^0-9]", "", hs)[:10]
    for pat, sctg in _HS_SCTG:
        if re.match(pat, hs):
            return sctg
    return "31"


def _parse_imdb_zip(zip_path: Path) -> list[dict[str, Any]]:
    """Parse IMP_COMM.txt rows from Census IMDB zip."""
    events: list[dict[str, Any]] = []
    with zipfile.ZipFile(zip_path) as zf:
        name = next((n for n in zf.namelist() if n.upper() == "IMP_COMM.TXT"), None)
        if not name:
            return []
        with zf.open(name) as fh:
            for raw in fh:
                line = raw.decode("utf-8", errors="ignore").rstrip("\n\r")
                if len(line) < 30:
                    continue
                hs = line[:10].strip()
                if not hs.isdigit():
                    continue
                m = re.search(r"(20\d{4})", line[10:])
                if not m:
                    continue
                month = m.group(1)
                sctg2 = _hs_to_sctg(hs)
                nums = [int(x) for x in re.findall(r"\b\d{4,}\b", line[60:])]
                weight = float(nums[0]) if nums else 1.0
                events.append({"hs": hs, "sctg2": sctg2, "month": month, "weight": weight})
    return events


def _resolve_imdb_zip() -> Path | None:
    p = resolve_path("census_hs_trade")
    candidates: list[Path] = []
    if p and p.stat().st_size >= 1_000_000:
        candidates.append(p)
    trade_dir = UNWRAPPED / "trade"
    for release_dir in sorted(trade_dir.glob("release=*"), reverse=True):
        candidates.extend(sorted(release_dir.glob("IMDB*.ZIP"), reverse=True))
    for z in candidates:
        if z.stat().st_size < 1_000_000:
            continue
        try:
            with zipfile.ZipFile(z) as zf:
                names = {n.upper() for n in zf.namelist()}
                if "IMP_COMM.TXT" in names:
                    return z
        except zipfile.BadZipFile:
            continue
    return None


def append_eia_events(
    con: duckdb.DuckDBPyConnection,
    *,
    input_files: list[str],
    warnings: list[str],
) -> int:
    count = 0
    try:
        pet = require_path("pet_imports")
        pet_rel = str(pet.relative_to(ROOT))
        pet_mtime = pet.stat().st_mtime
        input_files.append(pet_rel)

        rows: list[tuple[str, str, float]] = []
        with pet.open(encoding="utf-8") as fh:
            for line in fh:
                line = line.strip()
                if not line:
                    continue
                try:
                    obj = json.loads(line)
                except json.JSONDecodeError:
                    continue
                sid = obj.get("series_id", "")
                if sid != "PET_IMPORTS.WORLD-US-ALL.M":
                    continue
                for period, val in obj.get("data") or []:
                    if val is None:
                        continue
                    rows.append((str(period), "17", float(val)))

        if rows:
            con.execute(
                """
                CREATE OR REPLACE TABLE pet_pressure AS
                SELECT * FROM (VALUES {})
                AS t(event_month, sctg2, weight)
                """.format(",".join(f"('{m}','{s}',{w})" for m, s, w in rows[:500]))
            )
            con.execute(
                f"""
                INSERT INTO all_events
                SELECT
                  md5('commodity_pressure|pet|' || event_month) AS event_id,
                  'commodity_pressure' AS edge_type,
                  'commodity' AS src_type, sctg2 AS src_id,
                  'national' AS dst_type, 'US' AS dst_id,
                  event_month AS event_time, 'month' AS event_grain,
                  weight, 'strong' AS confidence,
                  'eia_pet_imports_world_monthly' AS inference_method,
                  'maritime' AS source_family,
                  ? AS source_path, ?::BIGINT AS source_mtime,
                  json_object('series', 'PET_IMPORTS.WORLD-US-ALL.M', 'units', 'thousand_barrels') AS attrs_json
                FROM pet_pressure
                """,
                [pet_rel, pet_mtime],
            )
            count += con.execute(
                "SELECT COUNT(*)::BIGINT FROM pet_pressure"
            ).fetchone()[0]
    except FileNotFoundError:
        warnings.append("eia_pet_missing")

    try:
        coal = require_path("eia_coal")
        coal_rel = str(coal.relative_to(ROOT))
        coal_mtime = coal.stat().st_mtime
        input_files.append(coal_rel)
        coal_rows: list[tuple[str, str, float]] = []
        with coal.open(encoding="utf-8") as fh:
            for line in fh:
                if "COAL.PRODUCTION." not in line or "-TOT.A" not in line:
                    continue
                try:
                    obj = json.loads(line)
                except json.JSONDecodeError:
                    continue
                geo = obj.get("iso3166") or ""
                if not geo.startswith("USA-"):
                    continue
                state = geo.split("-", 1)[-1]
                data = obj.get("data") or []
                if not data:
                    continue
                year, val = data[0]
                if val is None or float(val) <= 0:
                    continue
                coal_rows.append((str(year), state, float(val)))

        if coal_rows:
            con.execute(
                """
                CREATE OR REPLACE TABLE coal_pressure AS
                SELECT * FROM (VALUES {})
                AS t(event_year, state, weight)
                """.format(",".join(f"('{y}','{st}',{w})" for y, st, w in coal_rows[:2000]))
            )
            con.execute(
                f"""
                INSERT INTO all_events
                SELECT
                  md5('commodity_pressure|coal|' || state || '|' || event_year) AS event_id,
                  'commodity_pressure' AS edge_type,
                  'commodity' AS src_type, '19' AS src_id,
                  'state' AS dst_type, state AS dst_id,
                  event_year AS event_time, 'year' AS event_grain,
                  weight, 'medium' AS confidence,
                  'eia_coal_production_state_annual' AS inference_method,
                  'rail' AS source_family,
                  ? AS source_path, ?::BIGINT AS source_mtime,
                  json_object('sctg2', '19', 'state', state) AS attrs_json
                FROM coal_pressure
                """,
                [coal_rel, coal_mtime],
            )
            count += con.execute("SELECT COUNT(*)::BIGINT FROM coal_pressure").fetchone()[0]
    except FileNotFoundError:
        warnings.append("eia_coal_missing")

    return count


def append_trade_events(
    con: duckdb.DuckDBPyConnection,
    *,
    input_files: list[str],
    warnings: list[str],
) -> int:
    zip_path = _resolve_imdb_zip()
    if not zip_path:
        warnings.append("imdb_zip_missing")
        return 0

    zip_rel = str(zip_path.relative_to(ROOT))
    zip_mtime = zip_path.stat().st_mtime
    input_files.append(zip_rel)

    parsed = _parse_imdb_zip(zip_path)
    if not parsed:
        warnings.append("imdb_parse_empty")
        return 0

    latest_month = max(r["month"] for r in parsed)
    by_key: dict[tuple[str, str], float] = {}
    for row in parsed:
        if row["month"] != latest_month:
            continue
        key = (row["month"], row["sctg2"])
        by_key[key] = by_key.get(key, 0) + row["weight"]

    values = ",".join(f"('{m}','{s}',{w})" for (m, s), w in by_key.items())
    con.execute(
        f"""
        CREATE OR REPLACE TABLE trade_agg AS
        SELECT * FROM (VALUES {values}) AS t(event_month, sctg2, weight)
        """
    )
    con.execute(
        f"""
        INSERT INTO all_events
        SELECT
          md5('trade_import|' || sctg2 || '|' || event_month) AS event_id,
          'trade_import' AS edge_type,
          'national' AS src_type, 'US' AS src_id,
          'commodity' AS dst_type, sctg2 AS dst_id,
          event_month AS event_time, 'month' AS event_grain,
          weight, 'medium' AS confidence,
          'census_imdb_imp_comm' AS inference_method,
          'trade' AS source_family,
          ? AS source_path, ?::BIGINT AS source_mtime,
          json_object('sctg2', sctg2, 'imdb_month', event_month) AS attrs_json
        FROM trade_agg
        """,
        [zip_rel, zip_mtime],
    )
    return con.execute("SELECT COUNT(*)::BIGINT FROM trade_agg").fetchone()[0]


def append_carrier_risk_events(
    con: duckdb.DuckDBPyConnection,
    *,
    input_files: list[str],
    warnings: list[str],
) -> int:
    sms = resolve_family_file("fmcsa_carriers", "sms_input/sms_input_violation.csv", root_level=True)
    ins = resolve_family_file(
        "fmcsa_carriers", "insurance/actpendinsur_all_with_history.csv", root_level=True, min_bytes=1_000_000
    )
    if not sms or not ins:
        warnings.append("fmcsa_sms_insurance_missing")
        return 0

    sms_rel = str(sms.relative_to(ROOT))
    ins_rel = str(ins.relative_to(ROOT))
    input_files.extend([sms_rel, ins_rel])

    con.execute(
        f"""
        INSERT INTO all_events
        WITH viol AS (
          SELECT TRY_CAST(DOT_Number AS BIGINT) AS dot,
            SUBSTR(REPLACE(CAST(Insp_Date AS VARCHAR), '-', ''), 1, 6) AS event_month,
            COUNT(*)::DOUBLE AS viol_count,
            SUM(COALESCE(TRY_CAST(Total_Severity_Wght AS DOUBLE), 0)) AS sev
          FROM read_csv_auto(?, {CSV_OPTS})
          WHERE DOT_Number IS NOT NULL AND Insp_Date IS NOT NULL
          GROUP BY 1, 2
        ),
        ins AS (
          SELECT TRY_CAST(DOT_NUMBER AS BIGINT) AS dot,
            COUNT(*)::DOUBLE AS policies
          FROM read_csv_auto(?, {CSV_OPTS})
          WHERE DOT_NUMBER IS NOT NULL
          GROUP BY 1
        ),
        joined AS (
          SELECT v.dot, v.event_month, v.viol_count,
            v.viol_count + 0.1 * v.sev + COALESCE(i.policies, 0) AS risk_weight
          FROM viol v
          LEFT JOIN ins i ON v.dot = i.dot
        )
        SELECT
          md5('carrier_risk|' || CAST(dot AS VARCHAR) || '|' || event_month) AS event_id,
          'carrier_risk' AS edge_type,
          'carrier' AS src_type, CAST(dot AS VARCHAR) AS src_id,
          'national' AS dst_type, 'US' AS dst_id,
          event_month AS event_time, 'month' AS event_grain,
          risk_weight AS weight,
          CASE WHEN risk_weight > 50 THEN 'medium' ELSE 'weak' END AS confidence,
          'fmcsa_sms_violation_insurance' AS inference_method,
          'fmcsa_carriers' AS source_family,
          ? AS source_path, ?::BIGINT AS source_mtime,
          json_object('violations', viol_count) AS attrs_json
        FROM joined
        WHERE dot IS NOT NULL
        """,
        [str(sms), str(ins), sms_rel, sms.stat().st_mtime],
    )
    return con.execute(
        "SELECT COUNT(*)::BIGINT FROM all_events WHERE edge_type = 'carrier_risk'"
    ).fetchone()[0]


def append_pipeline_events(
    con: duckdb.DuckDBPyConnection,
    *,
    input_files: list[str],
    warnings: list[str],
) -> int:
    phmsa_root = UNWRAPPED / "phmsa_pipelines" / "gas_distribution_annual_2010_present"
    files = sorted(phmsa_root.glob("GD * 2025.csv"), reverse=True)
    if not files:
        warnings.append("phmsa_pipeline_missing")
        return 0

    paths = [str(f) for f in files[:60]]
    paths_sql = ",".join(f"'{p}'" for p in paths)
    first = files[0]
    first_rel = str(first.relative_to(ROOT))
    input_files.append(first_rel)

    con.execute(
        f"""
        INSERT INTO all_events
        WITH raw AS (
          SELECT STOP AS state, TRY_CAST(MMILES_TOTAL AS DOUBLE) AS miles
          FROM read_csv_auto([{paths_sql}], {CSV_OPTS}, union_by_name=true)
          WHERE STOP IS NOT NULL AND LENGTH(TRIM(STOP)) = 2
        ),
        agg AS (
          SELECT state, SUM(miles) AS mileage
          FROM raw
          WHERE miles IS NOT NULL AND miles > 0
          GROUP BY 1
        )
        SELECT
          md5('pipeline_signal|' || state || '|2025') AS event_id,
          'pipeline_signal' AS edge_type,
          'state' AS src_type, state AS src_id,
          'national' AS dst_type, 'US' AS dst_id,
          '2025' AS event_time, 'year' AS event_grain,
          mileage AS weight,
          'medium' AS confidence,
          'phmsa_gas_distribution_miles' AS inference_method,
          'phmsa_pipelines' AS source_family,
          ? AS source_path, ?::BIGINT AS source_mtime,
          json_object('state', state, 'report_year', 2025) AS attrs_json
        FROM agg
        """,
        [first_rel, first.stat().st_mtime],
    )
    return con.execute(
        "SELECT COUNT(*)::BIGINT FROM all_events WHERE edge_type = 'pipeline_signal'"
    ).fetchone()[0]


def append_fmc_maritime_events(
    con: duckdb.DuckDBPyConnection,
    *,
    input_files: list[str],
    warnings: list[str],
) -> int:
    """FMC OTI + VOCC active licenses → maritime_license edges."""
    oti = UNWRAPPED / "fmc_maritime" / "oti_licensed_active.csv"
    vocc = UNWRAPPED / "fmc_maritime" / "vocc_active.csv"
    paths = [("oti", oti), ("vocc", vocc)]
    rows: list[tuple[str, str, str, float]] = []
    for kind, path in paths:
        if not path.exists() or path.stat().st_size < 100:
            warnings.append(f"fmc_{kind}_missing")
            continue
        rel = str(path.relative_to(ROOT))
        input_files.append(rel)
        mtime = path.stat().st_mtime
        con.execute(
            f"""
            CREATE OR REPLACE TABLE fmc_{kind} AS
            SELECT
              TRIM(CAST("Organization No." AS VARCHAR)) AS org_no,
              TRIM(CAST("Legal Name" AS VARCHAR)) AS legal_name,
              '{kind}' AS license_kind,
              {mtime}::DOUBLE AS source_mtime,
              '{rel}' AS source_path
            FROM read_csv_auto('{path}', header=true, ignore_errors=true, all_varchar=true)
            WHERE "Organization No." IS NOT NULL
              AND TRIM(CAST("Organization No." AS VARCHAR)) != ''
            """
        )
        n = con.execute(f"SELECT COUNT(*)::BIGINT FROM fmc_{kind}").fetchone()[0]
        if n:
            con.execute(
                f"""
                INSERT INTO all_events
                SELECT
                  md5('maritime_license|' || license_kind || '|' || org_no) AS event_id,
                  'maritime_license' AS edge_type,
                  'maritime_org' AS src_type, org_no AS src_id,
                  'national' AS dst_type, 'US' AS dst_id,
                  '2026' AS event_time, 'year' AS event_grain,
                  1.0 AS weight, 'strong' AS confidence,
                  'fmc_' || license_kind || '_active' AS inference_method,
                  'fmc_maritime' AS source_family,
                  source_path, source_mtime::BIGINT AS source_mtime,
                  json_object(
                    'org_no', org_no,
                    'legal_name', legal_name,
                    'license_kind', license_kind
                  ) AS attrs_json
                FROM fmc_{kind}
                """
            )
            rows.append((kind, rel, str(n), float(n)))
    # ensure node type maritime_org is acceptable via attrs only; schema NODE_TYPES may lack it
    return int(sum(r[3] for r in rows))


def _r1_workbook_paths() -> list[tuple[str, Path]]:
    """Return (railroad_mark, xlsx_path) for R1 2025 filings on disk."""
    out: list[tuple[str, Path]] = []
    mapping = [
        ("BNSF", UNWRAPPED / "stb_rail" / "R1-BNSF-2025.xlsx"),
        ("NS", UNWRAPPED / "stb_rail" / "R1-NS-2025.xlsx"),
        ("GTC", UNWRAPPED / "stb_rail" / "R1-GTC-2025.xlsx"),
        ("SOO", UNWRAPPED / "stb_rail" / "R1-SOO-KCSR-2025.xlsx"),
        ("CSXT", UNWRAPPED / "stb_rail" / "R1-CSX-2025" / "2025 Sch 755 Final.xlsx"),
        ("UP", UNWRAPPED / "stb_rail" / "R1-UP-2025" / "2025 Sch 755 Final.xlsx"),
    ]
    # CSX/UP fuel + mileage may be separate files
    extras = [
        ("CSXT", "750", UNWRAPPED / "stb_rail" / "R1-CSX-2025" / "2025 Sch 750 Final.xlsx"),
        ("CSXT", "700", UNWRAPPED / "stb_rail" / "R1-CSX-2025" / "2025 Sch 700 Final.xlsx"),
        ("UP", "750", UNWRAPPED / "stb_rail" / "R1-UP-2025" / "2025 Sch 750 Final.xlsx"),
        ("UP", "700", UNWRAPPED / "stb_rail" / "R1-UP-2025" / "2025 Sch 700 Final.xlsx"),
    ]
    for mark, path in mapping:
        if path.exists() and path.stat().st_size > 10_000:
            out.append((mark, path))
    return out


def _parse_r1_metrics(path: Path) -> dict[str, float]:
    """Extract key operating metrics from an R1 workbook (Sch 755/750/700)."""
    import openpyxl

    metrics: dict[str, float] = {}
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheets = {s.lower(): s for s in wb.sheetnames}

    def sheet_named(*cands: str) -> str | None:
        # Prefer exact sheet name; never pick instruction sheets.
        for c in cands:
            if c.lower() in sheets:
                return sheets[c.lower()]
        for c in cands:
            cl = c.lower()
            for s in wb.sheetnames:
                sl = s.lower()
                if "instr" in sl:
                    continue
                if cl == sl or sl.startswith(cl + " ") or sl.endswith(" " + cl):
                    return s
                # SOO-style: "77 S755"
                if f"s{cl}" in sl.replace(" ", "") or sl.endswith(cl):
                    return s
        return None

    def freightish(nums: list[float]) -> float | None:
        # Line/cross-check cols are small ints; freight values are large.
        big = [n for n in nums if n >= 100]
        return max(big) if big else None

    s755 = sheet_named("755")
    if s755:
        for row in wb[s755].iter_rows(values_only=True):
            vals = [c for c in row if c is not None]
            text = " ".join(str(c) for c in vals if isinstance(c, str)).upper()
            nums = [float(c) for c in vals if isinstance(c, (int, float))]
            if not nums:
                continue
            if "MILES OF ROAD OPERATED" in text:
                v = freightish(nums)
                if v is not None:
                    metrics["miles_of_road"] = v
            elif "TOTAL TRAIN MILES" in text or "TOTAL ALL TRAINS" in text:
                v = freightish(nums)
                if v is not None:
                    metrics["train_miles"] = v
            elif "TOTAL ALL SERVICES" in text and ("LOCOMOTIVE" in text or "3-31" in text):
                v = freightish(nums)
                if v is not None:
                    metrics["locomotive_unit_miles"] = v

    s750 = sheet_named("750")
    if s750:
        for row in wb[s750].iter_rows(values_only=True):
            vals = [c for c in row if c is not None]
            text = " ".join(str(c) for c in vals if isinstance(c, str)).upper()
            nums = [float(c) for c in vals if isinstance(c, (int, float))]
            if nums and "FREIGHT" in text and "PASSENGER" not in text and "TOTAL" not in text:
                # first freight diesel gallons line
                if "diesel_freight_gallons" not in metrics and nums[-1] > 1_000_000:
                    metrics["diesel_freight_gallons"] = nums[-1]

    s700 = sheet_named("700")
    if s700:
        for row in wb[s700].iter_rows(values_only=True):
            vals = [c for c in row if c is not None]
            text = " ".join(str(c) for c in vals if isinstance(c, str)).upper()
            nums = [float(c) for c in vals if isinstance(c, (int, float))]
            if "GRAND TOTAL" in text and nums:
                v = freightish(nums)
                if v is not None:
                    metrics["miles_operated_total"] = v

    return metrics


def append_stb_r1_events(
    con: duckdb.DuckDBPyConnection,
    *,
    input_files: list[str],
    warnings: list[str],
) -> int:
    """STB R1 2025 operating statistics → rail_metric edges (annual vintage)."""
    paths = _r1_workbook_paths()
    if not paths:
        warnings.append("stb_r1_2025_missing")
        return 0

    # Also parse CSX/UP companion 750/700 if primary is 755-only
    companion = {
        "CSXT": [
            UNWRAPPED / "stb_rail" / "R1-CSX-2025" / "2025 Sch 750 Final.xlsx",
            UNWRAPPED / "stb_rail" / "R1-CSX-2025" / "2025 Sch 700 Final.xlsx",
        ],
        "UP": [
            UNWRAPPED / "stb_rail" / "R1-UP-2025" / "2025 Sch 750 Final.xlsx",
            UNWRAPPED / "stb_rail" / "R1-UP-2025" / "2025 Sch 700 Final.xlsx",
        ],
    }

    event_rows: list[tuple[str, str, float, str, str, int]] = []
    for mark, path in paths:
        try:
            metrics = _parse_r1_metrics(path)
            for extra in companion.get(mark, []):
                if extra.exists():
                    metrics.update({k: v for k, v in _parse_r1_metrics(extra).items() if k not in metrics})
                    input_files.append(str(extra.relative_to(ROOT)))
        except Exception as exc:  # noqa: BLE001
            warnings.append(f"stb_r1_parse_{mark}:{type(exc).__name__}")
            continue
        if not metrics:
            warnings.append(f"stb_r1_empty_{mark}")
            continue
        rel = str(path.relative_to(ROOT))
        input_files.append(rel)
        mtime = int(path.stat().st_mtime)
        for metric, weight in metrics.items():
            if weight is None or weight <= 0:
                continue
            event_rows.append((mark, metric, float(weight), rel, "2025", mtime))

    if not event_rows:
        warnings.append("stb_r1_no_metrics")
        return 0

    values = ",".join(
        f"('{rr}','{metric}',{w},'{path}','{year}',{mtime})"
        for rr, metric, w, path, year, mtime in event_rows
    )
    con.execute(
        f"""
        CREATE OR REPLACE TABLE r1_metrics AS
        SELECT * FROM (VALUES {values})
        AS t(railroad_mark, metric, weight, source_path, event_year, source_mtime)
        """
    )
    con.execute(
        """
        INSERT INTO all_events
        SELECT
          md5('rail_metric|r1|' || railroad_mark || '|' || metric || '|' || event_year) AS event_id,
          'rail_metric' AS edge_type,
          'rail_operator' AS src_type, railroad_mark AS src_id,
          'national' AS dst_type, metric AS dst_id,
          event_year AS event_time, 'year' AS event_grain,
          weight, 'strong' AS confidence,
          'stb_r1_2025' AS inference_method,
          'stb_rail' AS source_family,
          source_path, source_mtime,
          json_object(
            'railroad_mark', railroad_mark,
            'metric', metric,
            'r1_year', event_year,
            'filing', 'R1'
          ) AS attrs_json
        FROM r1_metrics
        """
    )
    return len(event_rows)
